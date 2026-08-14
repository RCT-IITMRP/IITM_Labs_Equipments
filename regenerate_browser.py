"""
IIT Madras Labs & Equipment Browser — Regeneration Script
=========================================================
Reads equipment data from an Excel workbook and generates a self-contained
HTML browser with search, filtering, card/table views, and an AI assistant.

Usage:
    python3 regenerate_browser.py

Requirements:
    pip install openpyxl

Place this script in the same folder as your Excel source file.
Update SOURCE_FILE below if the filename changes.

The workbook must contain a 'Departments_and_Entities' sheet with Department
and Entity Type columns used to populate filter dropdowns. Every other sheet
is treated as equipment data; columns are detected dynamically from headers.

The AI Equipment Assistant calls a Cloudflare Worker proxy that holds the
Gemini API key server-side — no secret is embedded in the generated HTML.
Configure ASSISTANT_PROXY_URL below. See DEPLOYMENT.md for setup details.
"""

import json
import os
import openpyxl

# ── Configuration ─────────────────────────────────────────────────────────────
SOURCE_FILE = "IIT-M_L&E_Web&Visit_Data_OG.xlsx"
OUTPUT_FILE = "index.html"

# Cloudflare Worker proxy URL for the AI Equipment Assistant.
# Leave empty to disable the assistant. See DEPLOYMENT.md for setup.
ASSISTANT_PROXY_URL = "https://iitm-equipment-assistant.dharman.workers.dev"

# Sheet used exclusively for populating filter dropdowns (not displayed as data).
LOOKUP_SHEET = 'Departments_and_Entities'


def clean(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)

    return str(v).strip().replace('\n', ' ').replace('\xa0', '').strip()


def extract_data(filepath):
    """Dynamically read every sheet except LOOKUP_SHEET.

    Returns (equipment_list, filter_options_dict).
    filter_options_dict = {'departments': [...], 'entity_types': [...]}.
    """
    import zipfile, os
    # Validate before handing to openpyxl so we get a clear error message
    # instead of a cryptic BadZipFile traceback.
    file_size = os.path.getsize(filepath)
    if file_size < 1024:
        try:
            with open(filepath, 'r', errors='replace') as fh:
                content = fh.read(500)
        except Exception:
            content = '<unreadable>'
        raise SystemExit(
            f"\nERROR: '{filepath}' is only {file_size} bytes — not a valid Excel file.\n"
            f"File content: {repr(content)}\n\n"
            "LIKELY CAUSE: The OneDrive URL returned an HTML error/redirect page\n"
            "instead of the raw .xlsx file. See the YAML workflow comment for\n"
            "how to obtain the correct direct-download URL."
        )
    if not zipfile.is_zipfile(filepath):
        raise SystemExit(
            f"\nERROR: '{filepath}' is not a valid ZIP/xlsx file ({file_size:,} bytes).\n"
            "Ensure the download URL points directly at the .xlsx binary, not an\n"
            "HTML viewer page."
        )

    wb = openpyxl.load_workbook(filepath, read_only=True)

    # ── Read filter options from the lookup sheet ───────────────────────────
    filter_departments = []
    filter_entity_types = []
    if LOOKUP_SHEET in wb.sheetnames:
        ws_lookup = wb[LOOKUP_SHEET]
        rows = list(ws_lookup.iter_rows(values_only=True))
        if rows:
            header = [clean(h).lower() if h else '' for h in rows[0]]
            dept_col = None
            etype_col = None
            for i, h in enumerate(header):
                if 'department' in h:
                    dept_col = i
                elif 'entity' in h and 'type' in h:
                    etype_col = i
            for row in rows[1:]:
                if dept_col is not None and dept_col < len(row):
                    v = clean(row[dept_col])
                    if v and v not in filter_departments:
                        filter_departments.append(v)
                if etype_col is not None and etype_col < len(row):
                    v = clean(row[etype_col])
                    if v and v not in filter_entity_types:
                        filter_entity_types.append(v)
    filter_departments.sort()
    filter_entity_types.sort()

    # ── Dynamically read every other sheet ──────────────────────────────────
    equipment = []
    data_sheets = [s for s in wb.sheetnames if s != LOOKUP_SHEET]

    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            # Header-only or empty sheet — skip
            continue

        # Build header map: column index → cleaned header name
        raw_header = rows[0]
        headers = []
        for h in raw_header:
            ch = clean(h)
            headers.append(ch if ch else None)

        for row in rows[1:]:
            # Build a dict from all non-empty columns
            entry = {}
            all_empty = True
            for i, val in enumerate(row):
                if i >= len(headers) or headers[i] is None:
                    continue
                cv = clean(val)
                if cv:
                    entry[headers[i]] = cv
                    all_empty = True  # just a flag reset; real check below
            if not entry:
                continue
            # Need at least something meaningful — skip truly blank rows
            # (rows where every cell was empty)
            equipment.append(entry)

    print(f"  Extracted {len(equipment)} equipment entries from {len(data_sheets)} data sheets")
    return equipment, {'departments': filter_departments, 'entity_types': filter_entity_types}


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IITM-RP Labs & Equipment</title>
<link rel="icon" type="image/png" href="iitmrp_logo.png">
<style>
  :root {
    --navy: #0A2747;
    --navy-mid: #0D3461;
    --gold: #C9953A;
    --gold-light: #E8B96A;
    --cream: #F7F5F0;
    --white: #FFFFFF;
    --text: #1C2B3A;
    --text-muted: #5A6A7A;
    --border: #DDD8CF;
    --card-bg: #FFFFFF;
    --tag-bg: #EEF3FA;
    --tag-text: #2A4A7F;
    --hover: #F0EDE8;
    --shadow: 0 2px 8px rgba(10,39,71,0.10);
    --shadow-lg: 0 8px 32px rgba(10,39,71,0.14);
    --radius: 10px;
    --radius-sm: 6px;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', system-ui, sans-serif; background: var(--cream); color: var(--text); min-height: 100vh; font-size: 14px; line-height: 1.5; }
  header { background: linear-gradient(135deg, var(--navy) 0%, var(--navy-mid) 100%); color: white; position: sticky; top: 0; z-index: 100; box-shadow: 0 2px 16px rgba(0,0,0,0.25); }
  .header-top { display: flex; align-items: center; gap: 18px; padding: 16px 28px; border-bottom: 1px solid rgba(201,149,58,0.30); }
  .logo-block { display: flex; align-items: center; gap: 14px; flex-shrink: 0; }

  /* Institution logos — side-by-side with a gold separator line */
  .logos-wrap { display: flex; align-items: center; gap: 10px; border-right: 1px solid rgba(201,149,58,0.35); padding-right: 14px; }
  .logo-img { height: 48px; width: auto; flex-shrink: 0; object-fit: contain; }

  .logo-text h1 { font-size: 17px; font-weight: 700; color: white; }
  .logo-text p { font-size: 11px; color: rgba(255,255,255,0.65); letter-spacing: 0.8px; text-transform: uppercase; }
  .header-stats { display: flex; gap: 24px; margin-left: auto; }
  .stat-pill { display: flex; flex-direction: column; align-items: center; background: rgba(255,255,255,0.08); border: 1px solid rgba(255,255,255,0.12); border-radius: 8px; padding: 6px 16px; }
  .stat-pill .num { font-size: 20px; font-weight: 800; color: var(--gold-light); }
  .stat-pill .lbl { font-size: 10px; text-transform: uppercase; letter-spacing: 0.8px; color: rgba(255,255,255,0.55); }
  /* Info bar — contact guidance shown between header-top and filter-bar */
  .header-info-bar { padding: 8px 28px; font-size: 13px; color: rgba(255,255,255,0.88); background: rgba(0,0,0,0.15); border-top: 1px solid rgba(201,149,58,0.20); border-bottom: 1px solid rgba(201,149,58,0.20); line-height: 1.55; }
  .header-info-bar a { color: var(--gold-light); text-decoration: underline; text-underline-offset: 2px; }
  .filter-bar { display: flex; align-items: center; gap: 10px; padding: 12px 28px; flex-wrap: wrap; background: rgba(0,0,0,0.12); }

  /* Shared search-wrap styles (used by both global and equipment-only search) */
  .search-wrap { position: relative; flex: 1; min-width: 200px; }
  .search-wrap svg { position: absolute; left: 12px; top: 50%; transform: translateY(-50%); color: rgba(255,255,255,0.5); pointer-events: none; }
  .search-wrap input { width: 100%; background: rgba(255,255,255,0.10); border: 1px solid rgba(255,255,255,0.18); border-radius: 8px; padding: 9px 12px 9px 38px; color: white; font-size: 13px; outline: none; }
  .search-wrap input::placeholder { color: rgba(255,255,255,0.45); }
  .search-wrap input:focus { border-color: var(--gold-light); background: rgba(255,255,255,0.15); }

  /* Equipment-name-only search bar — gold accent to distinguish from global search */
  .equip-only-wrap { flex: 0.85; min-width: 185px; }
  .equip-only-wrap input { border-color: rgba(201,149,58,0.40); }
  .equip-only-wrap input:focus { border-color: var(--gold-light); background: rgba(255,255,255,0.15); }
  .equip-only-wrap svg { color: rgba(201,149,58,0.80); }

  select { background: rgba(255,255,255,0.10); border: 1px solid rgba(255,255,255,0.18); border-radius: 8px; padding: 9px 32px 9px 12px; color: white; font-size: 12px; outline: none; cursor: pointer; appearance: none; -webkit-appearance: none; background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24'%3E%3Cpath fill='rgba(255,255,255,0.5)' d='M7 10l5 5 5-5z'/%3E%3C/svg%3E"); background-repeat: no-repeat; background-position: right 10px center; min-width: 160px; }
  select option { background: var(--navy); color: white; }
  .view-toggle { display: flex; background: rgba(255,255,255,0.10); border: 1px solid rgba(255,255,255,0.18); border-radius: 8px; overflow: hidden; }
  .view-btn { background: none; border: none; color: rgba(255,255,255,0.55); padding: 8px 12px; cursor: pointer; transition: all 0.2s; display: flex; align-items: center; }
  .view-btn.active { background: var(--gold); color: white; }
  .clear-btn { background: none; border: 1px solid rgba(255,255,255,0.20); color: rgba(255,255,255,0.65); border-radius: 8px; padding: 8px 14px; cursor: pointer; font-size: 12px; white-space: nowrap; }
  .clear-btn:hover { background: rgba(255,255,255,0.10); color: white; }
  main { padding: 20px 28px; }
  .results-meta { display: flex; align-items: center; margin-bottom: 16px; }
  .results-count { font-size: 13px; color: var(--text-muted); }
  .results-count strong { color: var(--navy); font-weight: 700; }

  /* Card-view: hidden by default, shown as grid when .active is set */
  #card-view { display: none; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap: 14px; }
  #card-view.active { display: grid; }

  .eq-card { background: var(--card-bg); border: 1px solid var(--border); border-radius: var(--radius); padding: 16px; box-shadow: var(--shadow); transition: transform 0.15s, box-shadow 0.15s; position: relative; overflow: hidden; }
  .eq-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px; background: var(--dept-color, var(--gold)); }
  .eq-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-lg); }
  .card-dept-tag { display: inline-block; font-size: 10px; font-weight: 600; letter-spacing: 0.6px; text-transform: uppercase; background: var(--tag-bg); color: var(--tag-text); border-radius: 4px; padding: 2px 8px; margin-bottom: 8px; }
  .card-equip-name { font-size: 15px; font-weight: 700; color: var(--navy); margin-bottom: 4px; line-height: 1.3; }
  .card-lab-name { font-size: 12px; color: var(--text-muted); margin-bottom: 10px; display: flex; align-items: flex-start; gap: 5px; }
  .card-lab-name svg { flex-shrink: 0; margin-top: 1px; }
  .card-divider { height: 1px; background: var(--border); margin: 10px 0; }
  .card-info-row { display: flex; align-items: flex-start; gap: 8px; font-size: 12px; color: var(--text-muted); margin-top: 5px; }
  .card-info-row svg { flex-shrink: 0; margin-top: 1px; }
  .card-info-row a { color: var(--gold); text-decoration: none; }
  .card-info-row .label { font-weight: 600; color: var(--text); min-width: 65px; }
 #table-view { display: none; }
  .table-scroll { overflow: auto; max-height: calc(100vh - var(--header-h, 130px) - 60px); border-radius: var(--radius); box-shadow: var(--shadow); }
  #table-view.active { display: block; }
  table { width: 100%; border-collapse: collapse; background: white; font-size: 13px; }
  thead { background: var(--navy); color: white; }
  th { padding: 12px 14px; text-align: left; font-size: 11px; font-weight: 600; letter-spacing: 0.7px; text-transform: uppercase; white-space: nowrap; cursor: pointer; user-select: none; }

  /* Sticky table header — sits below the sticky page header via --header-h */
  thead th { position: sticky; top: 0; z-index: 90; background: var(--navy); }

  tbody tr { border-bottom: 1px solid var(--border); transition: background 0.1s; }
  tbody tr:hover { background: var(--hover); }
  td { padding: 11px 14px; vertical-align: top; }
  td:first-child { font-weight: 600; color: var(--navy); }
  .dept-badge { display: inline-block; font-size: 10px; font-weight: 600; border-radius: 4px; padding: 2px 7px; white-space: nowrap; }
  .contact-link { color: var(--gold); text-decoration: none; font-size: 12px; }
  .dept-Physics{--dept-color:#8B5CF6}.dept-Aerospace-Engineering{--dept-color:#0EA5E9}.dept-Mechanical-Engineering{--dept-color:#F59E0B}.dept-Civil-Engineering{--dept-color:#10B981}.dept-Electrical-Engineering{--dept-color:#EF4444}.dept-Chemical-Engineering{--dept-color:#EC4899}.dept-Biotechnology{--dept-color:#14B8A6}.dept-Chemistry{--dept-color:#84CC16}.dept-Ocean-Engineering{--dept-color:#0284C7}.dept-Metallurgical{--dept-color:#78716C}.dept-Applied-Mechanics{--dept-color:#F97316}.dept-ARCI{--dept-color:#A855F7}.dept-IC-SR{--dept-color:#1D4ED8}.dept-Engineering-Design{--dept-color:#D946EF}.dept-default{--dept-color:var(--gold)}
  .badge-Physics{background:#EDE9FE;color:#5B21B6}.badge-Aerospace{background:#E0F2FE;color:#0369A1}.badge-Mechanical{background:#FEF3C7;color:#92400E}.badge-Civil{background:#D1FAE5;color:#065F46}.badge-Electrical{background:#FEE2E2;color:#991B1B}.badge-Chemical{background:#FCE7F3;color:#9D174D}.badge-Biotechnology{background:#CCFBF1;color:#0F766E}.badge-Chemistry{background:#ECFCCB;color:#365314}.badge-Ocean{background:#E0F2FE;color:#075985}.badge-Metal{background:#F5F5F4;color:#44403C}.badge-Applied{background:#FFEDD5;color:#9A3412}.badge-ARCI{background:#F3E8FF;color:#6B21A8}.badge-ICSR{background:#DBEAFE;color:#1E3A8A}.badge-ED{background:#FDF4FF;color:#86198F}.badge-default{background:#F3F4F6;color:#374151}
  .empty-state { text-align: center; padding: 80px 20px; color: var(--text-muted); }
  .empty-state h3 { font-size: 18px; color: var(--text); margin-bottom: 6px; }
  footer { text-align: center; padding: 24px; font-size: 12px; color: var(--text-muted); border-top: 1px solid var(--border); margin-top: 20px; }
  @media (max-width: 700px) { .header-stats{display:none} .header-top,.filter-bar{padding:12px 16px} main{padding:14px 16px} #card-view{grid-template-columns:1fr} }

  .header-info-bar a,
  .header-info-bar .link-color { color: var(--gold-light); font-weight: bold; }

  /* AI FAB — floating action button to open the assistant */
  #ai-fab { position: fixed; right: 24px; bottom: 24px; width: 56px; height: 56px; border-radius: 50%; background: linear-gradient(135deg, var(--gold) 0%, var(--gold-light) 100%); border: none; box-shadow: var(--shadow-lg); cursor: pointer; display: flex; align-items: center; justify-content: center; color: var(--navy); z-index: 300; transition: transform 0.15s; }
  #ai-fab:hover { transform: scale(1.06); }
  #ai-fab svg { width: 24px; height: 24px; }

  /* Backdrop overlay — dims the page behind the assistant panel */
  #ai-backdrop { position: fixed; inset: 0; background: rgba(0,0,0,0.50); z-index: 250;
    visibility: hidden; opacity: 0; pointer-events: none;
    transition: opacity 0.25s ease, visibility 0s linear 0.25s; }
  #ai-backdrop.show { visibility: visible; opacity: 1; pointer-events: auto;
    transition: opacity 0.25s ease, visibility 0s linear 0s; }

  /* AI panel — chat container. Uses top/left/width/height for resize support.
     Default position mirrors the original (right-aligned, below header).
     JS sets inline top/left/width/height on first open and during resize. */
  #ai-panel { position: fixed; right: 24px; top: calc(var(--header-h, 130px) + 0px); bottom: 85px; width: 450px; max-width: calc(100vw - 32px); background: var(--cream); border-radius: var(--radius); box-shadow: 0 12px 48px rgba(10,39,71,0.28); display: flex; flex-direction: column; overflow: visible; z-index: 300; border: 1px solid var(--border);
    visibility: hidden; opacity: 0; transform: scale(0.90) translateY(14px); transform-origin: bottom right; pointer-events: none;
    transition: opacity 0.25s ease, transform 0.25s ease, visibility 0s linear 0.25s;
    min-width: 300px; min-height: 280px; }
  #ai-panel.open { visibility: visible; opacity: 1; transform: scale(1) translateY(0); pointer-events: auto;
    transition: opacity 0.25s ease, transform 0.25s ease, visibility 0s linear 0s; }
  /* Suppress smooth transitions while the user is actively dragging a resize handle */
  #ai-panel.resizing { transition: none !important; }

  /* Resize handles — invisible grab zones on every edge and corner */
  .ai-resize { position: absolute; z-index: 310; }
  .ai-resize-n  { top: -4px;  left: 8px;  right: 8px; height: 8px; cursor: n-resize; }
  .ai-resize-s  { bottom: -4px; left: 8px; right: 8px; height: 8px; cursor: s-resize; }
  .ai-resize-e  { right: -4px; top: 8px; bottom: 8px; width: 8px; cursor: e-resize; }
  .ai-resize-w  { left: -4px;  top: 8px; bottom: 8px; width: 8px; cursor: w-resize; }
  .ai-resize-nw { top: -5px;  left: -5px;  width: 14px; height: 14px; cursor: nw-resize; }
  .ai-resize-ne { top: -5px;  right: -5px; width: 14px; height: 14px; cursor: ne-resize; }
  .ai-resize-sw { bottom: -5px; left: -5px; width: 14px; height: 14px; cursor: sw-resize; }
  .ai-resize-se { bottom: -5px; right: -5px; width: 14px; height: 14px; cursor: se-resize; }

  /* Panel inner sections */
  .ai-header { background: linear-gradient(135deg, var(--navy) 0%, var(--navy-mid) 100%); color: white; padding: 14px 16px; display: flex; align-items: center; gap: 10px; flex-shrink: 0; border-radius: var(--radius) var(--radius) 0 0; }
  .ai-header .ai-title { font-size: 14px; font-weight: 700; }
  .ai-header .ai-sub { font-size: 10px; color: rgba(255,255,255,0.6); margin-top: 1px; }
  .ai-header-text { flex: 1; }
  .ai-close { background: none; border: none; color: rgba(255,255,255,0.7); cursor: pointer; padding: 4px; display: flex; }
  .ai-close:hover { color: white; }
  .ai-messages { flex: 1; overflow-y: auto; padding: 14px; display: flex; flex-direction: column; gap: 10px; }
  .ai-msg { max-width: 88%; font-size: 13px; line-height: 1.45; padding: 9px 12px; border-radius: 12px; white-space: pre-wrap; }
  .ai-msg.user { align-self: flex-end; background: var(--navy); color: white; border-bottom-right-radius: 3px; }
  .ai-msg.bot { align-self: flex-start; background: white; color: var(--text); border: 1px solid var(--border); border-bottom-left-radius: 3px; }
  .ai-msg.error { align-self: flex-start; background: #FEF2F2; color: #991B1B; border: 1px solid #FECACA; }
  .ai-results { display: flex; flex-direction: column; gap: 8px; align-self: stretch; }
  .ai-result-card { background: white; border: 1px solid var(--border); border-radius: 8px; padding: 10px 12px; cursor: pointer; transition: border-color 0.15s, background 0.15s; }
  .ai-result-card:hover { border-color: var(--gold); background: var(--tag-bg); }
  .ai-result-eq { font-size: 13px; font-weight: 700; color: var(--navy); margin-bottom: 2px; }
  .ai-result-lab { font-size: 11px; color: var(--text-muted); margin-bottom: 4px; }
  .ai-result-why { font-size: 11px; color: #946715; font-style: italic; }
  .ai-typing { display: flex; gap: 4px; padding: 9px 12px; align-self: flex-start; }
  .ai-typing span { width: 6px; height: 6px; border-radius: 50%; background: var(--text-muted); opacity: 0.5; animation: ai-bounce 1.2s infinite; }
  .ai-typing span:nth-child(2) { animation-delay: 0.15s; }
  .ai-typing span:nth-child(3) { animation-delay: 0.3s; }
  @keyframes ai-bounce { 0%,60%,100%{transform:translateY(0);opacity:.5} 30%{transform:translateY(-4px);opacity:1} }
  .ai-input-row { display: flex; gap: 8px; padding: 12px; border-top: 1px solid var(--border); background: white; flex-shrink: 0; }
  .ai-input-row input { flex: 1; border: 1px solid var(--border); border-radius: 8px; padding: 9px 12px; font-size: 13px; outline: none; }
  .ai-input-row input:focus { border-color: var(--gold); }
  .ai-input-row button { background: var(--navy); color: white; border: none; border-radius: 8px; width: 38px; flex-shrink: 0; cursor: pointer; display: flex; align-items: center; justify-content: center; }
  .ai-input-row button:disabled { opacity: 0.5; cursor: default; }
  .ai-disclaimer { font-size: 12px; color: var(--text-muted); padding: 8px 14px 12px; text-align: center; flex-shrink: 0; line-height: 1.5; }
  .ai-disclaimer a { color: var(--gold); }

  @media (max-width: 700px) { #ai-panel { right: 12px; left: 12px; width: auto; top: calc(var(--header-h, 130px) + 6px); bottom: 80px; } #ai-fab { right: 16px; bottom: 16px; } }
</style>
</head>
<body>
<header>
  <div class="header-top">
    <div class="logo-block">

      <!-- Institution logos -->
      <div class="logos-wrap">
        <img src="iitm_logo.png" alt="IIT Madras" class="logo-img">
        <img src="iitmrp_logo.png" alt="IIT Madras Research Park" class="logo-img">
      </div>

      <div class="logo-text">
        <h1>IIT Madras/IITM Research Park Ecosystem(Beta) — Labs &amp; Equipments</h1>
        <p>Research Facilities Directory</p>
      </div>
    </div>
    <div class="header-stats">
      <div class="stat-pill"><span class="num" id="stat-total">-</span><span class="lbl">Equipment</span></div>
      <div class="stat-pill"><span class="num" id="stat-depts">-</span><span class="lbl">IITM Dept / IITMRP Clients</span></div>
      <div class="stat-pill"><span class="num" id="stat-labs">-</span><span class="lbl">Laboratories</span></div>
    </div>
  </div>
  <!-- Contact guidance info bar -->
  <div class="header-info-bar">
    For additional details, please reach out to us at <a href="mailto:rct@respark.iitm.ac.in">rct@respark.iitm.ac.in</a>. If PI name/number is available you can contact the PI. Please prefix <span class="link-color">2257</span> before the 4 digit extension number.
  </div>
  <div class="filter-bar">

    <!-- Global search: matches equipment name, lab, professor, department, operator -->
    <div class="search-wrap">
      <svg width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>
      <input type="text" id="search-input" placeholder="Search equipment, lab, or professor…" autocomplete="off">
    </div>

    <!-- Equipment-name-only search bar (gold accent, filters only equipment name) -->
    <div class="search-wrap equip-only-wrap">
      <svg width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/><path d="M8 11h6M11 8v6"/></svg>
      <input type="text" id="equip-search-input" placeholder="Search equipment name only…" autocomplete="off">
    </div>

    <select id="dept-filter"><option value="">IITM Dept / IITMRP Clients</option></select>
    <select id="entity-filter"><option value="">All Entity Types</option></select>

    <button class="clear-btn" id="clear-btn">✕ Clear</button>
    <div class="view-toggle">
      <button class="view-btn active" id="btn-card" title="Card view">
        <svg width="16" height="16" fill="currentColor" viewBox="0 0 24 24"><rect x="2" y="2" width="9" height="9" rx="1"/><rect x="13" y="2" width="9" height="9" rx="1"/><rect x="2" y="13" width="9" height="9" rx="1"/><rect x="13" y="13" width="9" height="9" rx="1"/></svg>
      </button>
      <button class="view-btn" id="btn-table" title="Table view">
        <svg width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M3 5h18M3 10h18M3 15h18M3 20h18M8 5v15M16 5v15"/></svg>
      </button>
    </div>
  </div>
</header>
<main>
  <div class="results-meta"><div class="results-count">Showing <strong id="result-count">0</strong> results</div></div>
  <div id="card-view" class="active"></div>
  <div id="table-view">
      <div class="table-scroll">
        <table>
          <thead id="table-head"><tr></tr></thead>
          <tbody id="table-body"></tbody>
        </table>
      </div>
    </div>
  <div class="empty-state" id="empty-state" style="display:none">
    <h3>No equipment found</h3><p>Try adjusting your search or filters.</p>
  </div>
</main>
<footer>IIT Madras Research Facilities Directory &nbsp;·&nbsp; Generated: __DATE__</footer>

<!-- AI Equipment Assistant — chat widget + backdrop overlay -->
<div id="ai-backdrop"></div>
<button id="ai-fab" title="Ask the Equipment Assistant" aria-label="Open equipment assistant">
  <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 11.5a8.38 8.38 0 0 1-.9 3.8 8.5 8.5 0 0 1-7.6 4.7 8.38 8.38 0 0 1-3.8-.9L3 21l1.9-5.7a8.38 8.38 0 0 1-.9-3.8 8.5 8.5 0 0 1 4.7-7.6 8.38 8.38 0 0 1 3.8-.9h.5a8.48 8.48 0 0 1 8 8v.5z"/></svg>
</button>
<div id="ai-panel">
  <!-- Resize handles: invisible grab zones on all 4 edges and 4 corners -->
  <div class="ai-resize ai-resize-n"  data-resize="n"></div>
  <div class="ai-resize ai-resize-s"  data-resize="s"></div>
  <div class="ai-resize ai-resize-e"  data-resize="e"></div>
  <div class="ai-resize ai-resize-w"  data-resize="w"></div>
  <div class="ai-resize ai-resize-nw" data-resize="nw"></div>
  <div class="ai-resize ai-resize-ne" data-resize="ne"></div>
  <div class="ai-resize ai-resize-sw" data-resize="sw"></div>
  <div class="ai-resize ai-resize-se" data-resize="se"></div>

  <div class="ai-header">
    <div class="ai-header-text">
      <div class="ai-title">AI-Powered Equipment Suggestion Assistant</div>
      <div class="ai-sub">Describe your research need</div>
    </div>
    <button class="ai-close" id="ai-close" aria-label="Close assistant">
      <svg width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><line x1="18" y1="6" x2="6" y2="18"/><line x1="6" y1="6" x2="18" y2="18"/></svg>
    </button>
  </div>
  <div class="ai-messages" id="ai-messages"></div>
  <div class="ai-input-row">
    <input type="text" id="ai-input" placeholder="e.g. measuring thermal conductivity of a nanofluid" autocomplete="off">
    <button id="ai-send" aria-label="Send">
      <svg width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><line x1="22" y1="2" x2="11" y2="13"/><polygon points="22 2 15 22 11 13 2 9 22 2"/></svg>
    </button>
  </div>
  <div class="ai-disclaimer">AI-generated suggestions — please confirm availability with the listed contact or the Research Collaboration Team at <a href="mailto:rct@respark.iitm.ac.in">rct@respark.iitm.ac.in</a>.</div>
</div>

<script>
/* ============================================================================
   DATA & CONFIGURATION
   ============================================================================ */

/** All equipment entries loaded from the workbook (array of dynamic-key objects). */
const equipmentData = __DATA__;

/** Filter dropdown options read from the Departments_and_Entities sheet. */
const filterOptions = __FILTER_OPTIONS__;

/** URL of the Cloudflare Worker proxy that holds the Gemini API key server-side.
 *  This URL is NOT a secret — safe to commit publicly. Leave empty to disable. */
const ASSISTANT_PROXY_URL = "__ASSISTANT_PROXY_URL__";


/* ============================================================================
   DYNAMIC FIELD ACCESSORS
   Column headers vary across sheets (e.g. "Lab Name" vs "Name of subordinate
   laboratories"). These helpers search common variations so the rest of the
   code can work regardless of which sheet an entry came from.
   ============================================================================ */

/** Return the first truthy value from `e` matching any key in `keys`. */
function eqField(e, keys) {
  for (const k of keys) { const v = e[k]; if (v) return v; }
  return '';
}
function eqName(e)       { return eqField(e, ['Equipment Name','equipment','Name of Equipment','Facilities']); }
function eqLab(e)        { return eqField(e, ['Lab Name','lab','Name of subordinate laboratories']); }
function eqProf(e)       { return eqField(e, ['PI','professor','Professor Incharge','Centre']); }
function eqDept(e)       { return eqField(e, ['Department','department','Deparment']); }
function eqEntityType(e) { return eqField(e, ['Entity Type','entity_type']); }
function eqOperator(e)   { return eqField(e, ['Operator Incharge','operator']); }
function eqOpEmail(e)    { return eqField(e, ['Operator Mail ID','op_email']); }
function eqContact(e)    { return eqField(e, ['Contact / Ext No','contact','Contact']); }
function eqProfEmail(e)  { return eqField(e, ['PI Email','prof_email']); }


/* ============================================================================
   AI CANDIDATE INDEX  (compact string sent to the LLM once per session)
   Format per line: index|equipment|lab|department|professor
   ============================================================================ */

const aiCandidateIndex = equipmentData.map((e, i) => {
  const trunc = (s, n) => { s = (s || '').trim(); return s.length > n ? s.slice(0, n) + '…' : s; };
  const prof = (eqProf(e) || '').replace(/\s*\([^)]*\)/g, '').trim();
  return `${i}|${trunc(eqName(e),90)}|${trunc(eqLab(e),70)}|${eqDept(e)}|${trunc(prof,40)}`;
}).join('\n');


/* ============================================================================
   DEPARTMENT COLOUR MAPPING
   Maps department keywords → CSS class pair for card accent + badge colours.
   ============================================================================ */

function deptClass(dept) {
  const d = (dept || '').toLowerCase();
  if (d.includes('physics'))             return { card:'dept-Physics',                badge:'badge-Physics' };
  if (d.includes('aerospace'))           return { card:'dept-Aerospace-Engineering',  badge:'badge-Aerospace' };
  if (d.includes('mechanical'))          return { card:'dept-Mechanical-Engineering', badge:'badge-Mechanical' };
  if (d.includes('civil'))               return { card:'dept-Civil-Engineering',      badge:'badge-Civil' };
  if (d.includes('electrical'))          return { card:'dept-Electrical-Engineering', badge:'badge-Electrical' };
  if (d.includes('chemical'))            return { card:'dept-Chemical-Engineering',   badge:'badge-Chemical' };
  if (d.includes('biotech'))             return { card:'dept-Biotechnology',          badge:'badge-Biotechnology' };
  if (d.includes('chemistry'))           return { card:'dept-Chemistry',              badge:'badge-Chemistry' };
  if (d.includes('ocean'))               return { card:'dept-Ocean-Engineering',      badge:'badge-Ocean' };
  if (d.includes('metallurg') || d.includes('material')) return { card:'dept-Metallurgical', badge:'badge-Metal' };
  if (d.includes('applied'))             return { card:'dept-Applied-Mechanics',      badge:'badge-Applied' };
  if (d.includes('arci'))                return { card:'dept-ARCI',                   badge:'badge-ARCI' };
  if (d.includes('ic') || d.includes('icsr')) return { card:'dept-IC-SR',             badge:'badge-ICSR' };
  if (d.includes('engineering design'))  return { card:'dept-Engineering-Design',     badge:'badge-ED' };
  return { card:'dept-default', badge:'badge-default' };
}


/* ============================================================================
   FILTER DROPDOWNS  — populated from the Departments_and_Entities lookup sheet
   ============================================================================ */

const deptSel   = document.getElementById('dept-filter');
const entitySel = document.getElementById('entity-filter');

filterOptions.departments.forEach(d => {
  const o = document.createElement('option'); o.value = d; o.textContent = d;
  deptSel.appendChild(o);
});
filterOptions.entity_types.forEach(d => {
  const o = document.createElement('option'); o.value = d; o.textContent = d;
  entitySel.appendChild(o);
});


/* ============================================================================
   FILTER / SEARCH / SORT STATE
   ============================================================================ */

let viewMode = 'card';
let sortCol  = -1;
let sortDir  = 1;
let currentData = [];

/** Collect current filter + search values from the UI controls. */
function getF() {
  return {
    q:          document.getElementById('search-input').value.trim().toLowerCase(),
    eq:         document.getElementById('equip-search-input').value.trim().toLowerCase(),
    dept:       deptSel.value,
    entityType: entitySel.value,
    lab:        ''   // lab filter is disabled; kept for future re-enablement
  };
}

/** Return the subset of equipmentData matching all active filters/searches. */
function filterData() {
  const { q, eq, dept, entityType, lab } = getF();
  return equipmentData.filter(e => {
    if (dept       && eqDept(e)       !== dept)       return false;
    if (entityType && eqEntityType(e) !== entityType) return false;
    if (lab        && eqLab(e)        !== lab)        return false;
    if (q) {
      const h = Object.values(e).join(' ').toLowerCase();
      if (!h.includes(q)) return false;
    }
    if (eq && !eqName(e).toLowerCase().includes(eq))  return false;
    return true;
  });
}


/* ============================================================================
   HTML ESCAPING & SEARCH-TERM HIGHLIGHTING
   ============================================================================ */

function esc(s) {
  return String(s || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

/** Escape `t`, then wrap every occurrence of `q` in a <mark> highlight. */
function hl(t, q) {
  if (!q || !t) return esc(t || '');
  const re = new RegExp('(' + q.replace(/[.*+?^${}()|[\]\\]/g,'\\$&') + ')','gi');
  return esc(t).replace(re, '<mark style="background:#FEF08A;padding:0 1px;border-radius:2px">$1</mark>');
}


/* ============================================================================
   CARD VIEW RENDERING
   Cards show well-known fields (equipment name, lab, PI, operator, contact) in
   a structured layout, then dynamically append any extra columns found in the
   row as additional info rows.
   ============================================================================ */

/** Set of column-header names that are rendered in the structured card layout.
 *  Extra keys NOT in this set get appended as generic info rows at the bottom. */
const CARD_TOP_KEYS_SET = new Set();
(function buildTopSet() {
  const groups = {
    equip:    ['Equipment Name','equipment','Name of Equipment','Facilities'],
    lab:      ['Lab Name','lab','Name of subordinate laboratories'],
    dept:     ['Department','department','Deparment'],
    prof:     ['PI','professor','Professor Incharge','Centre'],
    profEmail:['PI Email','prof_email'],
    operator: ['Operator Incharge','operator'],
    opEmail:  ['Operator Mail ID','op_email'],
    contact:  ['Contact / Ext No','contact','Contact']
  };
  Object.values(groups).forEach(arr => arr.forEach(k => CARD_TOP_KEYS_SET.add(k)));
})();

/** SVG icon fragments reused across card info rows. */
const SVG_HOME  = '<svg width="12" height="12" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"/><polyline points="9 22 9 12 15 12 15 22"/></svg>';
const SVG_USER  = '<svg width="13" height="13" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"/><circle cx="12" cy="7" r="4"/></svg>';
const SVG_PHONE = '<svg width="13" height="13" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M22 16.92v3a2 2 0 0 1-2.18 2 19.79 19.79 0 0 1-8.63-3.07A19.5 19.5 0 0 1 4.69 11.6 19.79 19.79 0 0 1 1.6 3.08 2 2 0 0 1 3.56 1h3a2 2 0 0 1 2 1.72c.127.96.361 1.903.7 2.81a2 2 0 0 1-.45 2.11L7.91 8.69a16 16 0 0 0 5.89 5.89l.9-.9a2 2 0 0 1 2.11-.45c.907.339 1.85.573 2.81.7A2 2 0 0 1 22 16.92z"/></svg>';
const SVG_PLUS  = '<svg width="13" height="13" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><circle cx="12" cy="12" r="10"/><line x1="12" y1="8" x2="12" y2="16"/><line x1="8" y1="12" x2="16" y2="12"/></svg>';

function renderCards(data, q) {
  const w = document.getElementById('card-view');
  if (!data.length) { w.innerHTML = ''; return; }

  w.innerHTML = data.map(e => {
    const dept  = eqDept(e);
    const cls   = deptClass(dept);
    const pd    = eqProf(e) || '';
    const em    = pd.match(/\(([^)]+@[^)]+)\)/);
    const email = em ? em[1].trim() : (eqProfEmail(e) || '');
    const pn    = pd.replace(/\s*\([^)]+\)/g, '').trim();
    const op    = eqOperator(e);
    const opEm  = eqOpEmail(e);
    const ct    = eqContact(e);

    // Extra fields: any column NOT already rendered in the structured layout
    const extraHtml = Object.keys(e)
      .filter(k => !CARD_TOP_KEYS_SET.has(k))
      .map(k => {
        const v = e[k];
        if (!v || !v.trim()) return '';
        return `<div class="card-info-row">${SVG_PLUS}<span><span class="label">${esc(k)}:</span> ${hl(v,q)}</span></div>`;
      }).join('');

    return `<div class="eq-card ${cls.card}">` +
      `<span class="card-dept-tag">${esc(dept || 'Unknown')}</span>` +
      `<div class="card-equip-name">${hl(eqName(e), q)}</div>` +
      `<div class="card-lab-name">${SVG_HOME}${hl(eqLab(e), q)}</div>` +
      ((pn || email || op || ct || extraHtml) ? '<div class="card-divider"></div>' : '') +
      (pn ? `<div class="card-info-row">${SVG_USER}<span><span class="label">PI:</span> ${hl(pn,q)}${email ? ` &nbsp;<a href="mailto:${esc(email)}" class="contact-link">${esc(email)}</a>` : ''}</span></div>` : '') +
      (op ? `<div class="card-info-row">${SVG_USER}<span><span class="label">Operator:</span> ${hl(op,q)}${opEm ? ` &nbsp;<a href="mailto:${esc(opEm)}" class="contact-link">${esc(opEm)}</a>` : ''}</span></div>` : '') +
      (ct && ct !== ' ' ? `<div class="card-info-row">${SVG_PHONE}<span><span class="label">Contact:</span> ${esc(ct)}</span></div>` : '') +
      extraHtml +
      `</div>`;
  }).join('');
}


/* ============================================================================
   TABLE VIEW RENDERING
   Columns are discovered dynamically from the union of all keys in the data.
   ============================================================================ */

const allColumnKeys = (function() {
  const s = new Set();
  equipmentData.forEach(e => Object.keys(e).forEach(k => s.add(k)));
  return [...s];
})();

function renderTable(data, q) {
  const th = document.getElementById('table-head');
  const tb = document.getElementById('table-body');
  th.innerHTML = '<tr>' + allColumnKeys.map((k, i) =>
    `<th onclick="sortTable(${i})">${esc(k)} <span class="sort-icon">↕</span></th>`
  ).join('') + '</tr>'; 
  tb.innerHTML = data.map(e =>
    '<tr>' + allColumnKeys.map(k => `<td>${hl(e[k] || '', q)}</td>`).join('') + '</tr>'
  ).join('');
}


/* ============================================================================
   RENDER ORCHESTRATION
   ============================================================================ */

/** Filter data, update count, then render the active view (card or table). */
function render() {
  const data = filterData();
  currentData = data;
  const { q, eq } = getF();
  const hlq = q || eq;
  document.getElementById('result-count').textContent = data.length.toLocaleString();
  document.getElementById('empty-state').style.display = data.length ? 'none' : 'block';
  if (viewMode === 'card') renderCards(data, hlq);
  else renderTable(data, hlq);
}

/** Sort the table by column index `col` and re-render. */
function sortTable(col) {
  if (sortCol === col) sortDir *= -1;
  else { sortCol = col; sortDir = 1; }
  document.querySelectorAll('#table-head th').forEach((th, i) => {
    th.classList.toggle('sorted', i === col);
    const ic = th.querySelector('.sort-icon');
    if (ic) ic.textContent = i === col ? (sortDir === 1 ? '↑' : '↓') : '↕';
  });
  const key = allColumnKeys[col];
  currentData.sort((a, b) => {
    const av = (a[key] || '').toLowerCase();
    const bv = (b[key] || '').toLowerCase();
    return av < bv ? -sortDir : av > bv ? sortDir : 0;
  });
  const { q, eq } = getF();
  renderTable(currentData, q || eq);
}


/* ============================================================================
   EVENT LISTENERS — search, filters, view toggle, clear
   ============================================================================ */

let debounceTimer;

document.getElementById('search-input').addEventListener('input', () => {
  clearTimeout(debounceTimer); debounceTimer = setTimeout(render, 180);
});
document.getElementById('equip-search-input').addEventListener('input', () => {
  clearTimeout(debounceTimer); debounceTimer = setTimeout(render, 180);
});
document.getElementById('dept-filter').addEventListener('change', render);
document.getElementById('entity-filter').addEventListener('change', render);

document.getElementById('clear-btn').addEventListener('click', () => {
  document.getElementById('search-input').value = '';
  document.getElementById('equip-search-input').value = '';
  deptSel.value = '';
  entitySel.value = '';
  render();
});

document.getElementById('btn-card').addEventListener('click', () => {
  viewMode = 'card';
  document.getElementById('btn-card').classList.add('active');
  document.getElementById('btn-table').classList.remove('active');
  document.getElementById('card-view').classList.add('active');
  document.getElementById('table-view').classList.remove('active');
  const { q, eq } = getF();
  renderCards(currentData, q || eq);
});
document.getElementById('btn-table').addEventListener('click', () => {
  viewMode = 'table';
  document.getElementById('btn-table').classList.add('active');
  document.getElementById('btn-card').classList.remove('active');
  document.getElementById('table-view').classList.add('active');
  document.getElementById('card-view').classList.remove('active');
  const { q, eq } = getF();
  renderTable(currentData, q || eq);
});


/* ============================================================================
   HEADER STATS & STICKY HEADER HEIGHT
   ============================================================================ */

document.getElementById('stat-total').textContent = equipmentData.length.toLocaleString();
document.getElementById('stat-depts').textContent = filterOptions.departments.length;
document.getElementById('stat-labs').textContent = [...new Set(equipmentData.map(e => eqLab(e)).filter(Boolean))].length;

/** Keep --header-h CSS variable in sync with actual header height (for sticky table header). */
function updateHeaderHeight() {
  const h = document.querySelector('header');
  if (h) document.documentElement.style.setProperty('--header-h', h.offsetHeight + 'px');
}
updateHeaderHeight();
if (window.ResizeObserver) { new ResizeObserver(updateHeaderHeight).observe(document.querySelector('header')); }
else { window.addEventListener('resize', updateHeaderHeight); }


/* ============================================================================
   AI EQUIPMENT ASSISTANT — Chat Logic
   Calls a Cloudflare Worker proxy (no API key in this file).
   ============================================================================ */

const AI_SYSTEM_PROMPT = `You are an equipment-matching assistant for the IIT Madras Research Park Labs & Equipment Directory. The user describes a research idea, technique, or measurement need in their own words — it may be vague, casual, or use different terminology than the equipment names. Your job is to infer the underlying technique or instrument family and select the most relevant entries from the CANDIDATE LIST given to you (format per line: index|equipment|lab|department|professor).

Rules:
- Only use indices that literally appear in the candidate list. Never invent equipment, labs, or indices.
- Rank by genuine relevance to the stated need, not by surface keyword overlap.
- Return at most 8 matches. Return fewer, or none, if fewer are genuinely relevant — do not pad the list.
- If nothing fits, return an empty "matches" array and use "reply" to suggest what kind of facility might help, or ask one clarifying question.
- "reply" must be 1-2 short, conversational sentences. No markdown, no headers, no lists inside "reply".
- "why" per match must be <=12 words, plain language.
- Respond with ONLY valid JSON, exactly this shape, no code fences, no extra text:
{"reply":"string","matches":[{"i":0,"why":"string"}]}`;

let aiHistory        = [];     // conversation turns: [{role:'user'|'model', parts:[{text}]}]
let aiCandidatesSent = false;  // candidate list is prepended to the first user message only
let aiBusy           = false;  // prevents overlapping API calls

/* Session flags — control which greeting/prompt is shown on open:
 *   aiWelcomeShown       — true after the initial welcome message is rendered
 *   aiHasInteracted      — true after the user sends their first message
 *   aiContinuationShown  — true after one continuation prompt has been shown;
 *                           prevents duplicates on repeated close/reopen cycles */
let aiWelcomeShown       = false;
let aiHasInteracted      = false;
let aiContinuationShown  = false;

const AI_CONTINUATION_PROMPTS = [
  "Welcome back — feel free to refine your previous search or explore a new requirement.",
  "Your prior conversation is intact. Is there a follow-up query I can assist with?",
  "Ready to continue. Would you like to narrow the results further or try a different research need?",
  "Happy to help further — let me know if you'd like to adjust your criteria or explore another area."
];


/* ── Panel open / close ───────────────────────────────────────────────────── */

function aiOpen() {
  const panel = document.getElementById('ai-panel');
  document.getElementById('ai-backdrop').classList.add('show');
  panel.classList.add('open');

  if (!aiWelcomeShown) {
    // First open of the page session — show the welcome greeting once.
    aiAppendBot("Hi! I\u2019m your IITMRP Equipment Suggestion Bot. Describe what you\u2019re trying to build, test, or measure and I\u2019ll find the most relevant labs and equipment for you.");
    aiWelcomeShown = true;
  } else if (aiHasInteracted && !aiContinuationShown) {
    // First reopen after the user has chatted — show one continuation prompt.
    const prompt = AI_CONTINUATION_PROMPTS[Math.floor(Math.random() * AI_CONTINUATION_PROMPTS.length)];
    aiAppendBot(prompt);
    aiContinuationShown = true;
  }
  // Subsequent reopens: no additional message appended.

  document.getElementById('ai-input').focus();
}

function aiClose() {
  document.getElementById('ai-panel').classList.remove('open');
  document.getElementById('ai-backdrop').classList.remove('show');
}


/* ── Chat message helpers ─────────────────────────────────────────────────── */

function aiAppendUser(text) {
  const w = document.getElementById('ai-messages');
  const d = document.createElement('div');
  d.className = 'ai-msg user';
  d.textContent = text;
  w.appendChild(d);
  w.scrollTop = w.scrollHeight;
}

function aiAppendBot(text, isError) {
  const w = document.getElementById('ai-messages');
  const d = document.createElement('div');
  d.className = 'ai-msg bot' + (isError ? ' error' : '');
  d.textContent = text;
  w.appendChild(d);
  w.scrollTop = w.scrollHeight;
}

function aiAppendResults(matches) {
  const w    = document.getElementById('ai-messages');
  const wrap = document.createElement('div');
  wrap.className = 'ai-results';
  matches.forEach(m => {
    const e = equipmentData[m.i];
    if (!e) return;
    const card = document.createElement('div');
    card.className = 'ai-result-card';
    card.innerHTML = `<div class="ai-result-eq">${esc(eqName(e))}</div>` +
      `<div class="ai-result-lab">${esc(eqLab(e))} · ${esc(eqDept(e))}</div>` +
      (m.why ? `<div class="ai-result-why">${esc(m.why)}</div>` : '');
    card.addEventListener('click', () => aiJumpTo(e));
    wrap.appendChild(card);
  });
  if (wrap.children.length) { w.appendChild(wrap); w.scrollTop = w.scrollHeight; }
}

/** Click an AI result card → fill the main search bar with the equipment name
 *  and scroll to the main results area. */
function aiJumpTo(e) {
  document.getElementById('equip-search-input').value = '';
  document.getElementById('search-input').value = eqName(e);
  deptSel.value = '';
  render();
  aiClose();
  document.querySelector('main').scrollIntoView({ behavior: 'smooth' });
}


/* ── Typing indicator ─────────────────────────────────────────────────────── */

function aiTypingShow() {
  const w = document.getElementById('ai-messages');
  const d = document.createElement('div');
  d.className = 'ai-typing'; d.id = 'ai-typing';
  d.innerHTML = '<span></span><span></span><span></span>';
  w.appendChild(d);
  w.scrollTop = w.scrollHeight;
}
function aiTypingHide() {
  const t = document.getElementById('ai-typing');
  if (t) t.remove();
}


/* ── Send message to Gemini via proxy ─────────────────────────────────────── */

async function aiSend() {
  if (aiBusy) return;
  const input = document.getElementById('ai-input');
  const text  = input.value.trim();
  if (!text) return;

  if (!ASSISTANT_PROXY_URL) {
    aiAppendUser(text);
    input.value = '';
    aiAppendBot("The assistant isn't configured yet — deploy the proxy (see DEPLOYMENT.md) and set ASSISTANT_PROXY_URL in regenerate_browser.py, then regenerate this page.", true);
    return;
  }

  aiBusy = true;
  document.getElementById('ai-send').disabled = true;
  aiHasInteracted = true;
  aiAppendUser(text);
  input.value = '';

  // On the first message, prepend the full candidate index for context
  const includeCandidates = !aiCandidatesSent;
  const userMessage = includeCandidates
    ? `CANDIDATE LIST:\n${aiCandidateIndex}\n\nUSER REQUEST: ${text}`
    : text;
  aiHistory.push({ role: 'user', parts: [{ text: userMessage }] });
  aiTypingShow();

  try {
    const res = await fetch(ASSISTANT_PROXY_URL, {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        systemInstruction: { parts: [{ text: AI_SYSTEM_PROMPT }] },
        contents: aiHistory,
        generationConfig: { responseMimeType: 'application/json', temperature: 0.3, thinkingConfig: { thinkingBudget: 0 } }
      })
    });
    aiTypingHide();

    if (!res.ok) {
      let msg = 'Something went wrong reaching the assistant. Please try again.';
      if      (res.status === 429) msg = 'The assistant has hit its free-tier usage limit for now — please try again in a minute.';
      else if (res.status === 403) msg = "The assistant proxy rejected this request (origin not allowed). If you just deployed it, check ALLOWED_ORIGINS in worker.js matches this page's URL.";
      else if (res.status === 400) msg = 'The assistant proxy is reachable but the request was rejected — check the Worker logs.';
      aiAppendBot(msg, true);
      aiHistory.pop();
      return;
    }

    const data = await res.json();
    // Gemini 2.5 Flash may return thought parts — skip those, find the real output
    const allParts  = data?.candidates?.[0]?.content?.parts || [];
    const outputPart = allParts.find(p => !p.thought) || allParts[0] || {};
    const raw = outputPart.text || '';
    aiHistory.push({ role: 'model', parts: [{ text: raw }] });
    if (includeCandidates) aiCandidatesSent = true;

    // Trim history to avoid unbounded growth
    const MAX_HISTORY_TURNS = 10;
    if (aiHistory.length > MAX_HISTORY_TURNS * 2) {
      aiHistory = aiHistory.slice(-(MAX_HISTORY_TURNS * 2));
    }

    let parsed;
    try {
      parsed = JSON.parse(raw.replace(/```json|```/g, '').trim());
    } catch (_) {
      aiAppendBot(raw || "I couldn't parse a response — please try rephrasing.", true);
      return;
    }
    aiAppendBot(parsed.reply || "Here's what I found:");
    const matches = (parsed.matches || [])
      .filter(m => Number.isInteger(m.i) && m.i >= 0 && m.i < equipmentData.length)
      .slice(0, 8);
    if (matches.length) aiAppendResults(matches);

  } catch (_) {
    aiTypingHide();
    aiAppendBot('Network error reaching the assistant. Check your connection and try again.', true);
    aiHistory.pop();
  } finally {
    aiBusy = false;
    document.getElementById('ai-send').disabled = false;
    input.focus();
  }
}


/* ============================================================================
   AI PANEL RESIZE  — drag from any edge or corner
   ============================================================================ */

(function initResize() {
  const panel   = document.getElementById('ai-panel');
  const MIN_W   = 300;
  const MIN_H   = 280;
  let active    = false;   // true while a resize drag is in progress
  let direction = '';      // e.g. 'n', 'se', 'w'
  let startX, startY;     // mouse position at drag start
  let startRect;           // panel's DOMRect at drag start

  /** Begin a resize drag when the user presses on a resize handle. */
  function onMouseDown(ev) {
    direction = ev.target.dataset.resize;
    if (!direction) return;
    ev.preventDefault();
    active    = true;
    startX    = ev.clientX;
    startY    = ev.clientY;
    startRect = panel.getBoundingClientRect();
    panel.classList.add('resizing');
    document.body.style.cursor = getComputedStyle(ev.target).cursor;
    document.body.style.userSelect = 'none';
  }

  /** Update panel geometry on every mousemove while dragging. */
  function onMouseMove(ev) {
    if (!active) return;
    const dx = ev.clientX - startX;
    const dy = ev.clientY - startY;

    let newTop    = startRect.top;
    let newLeft   = startRect.left;
    let newWidth  = startRect.width;
    let newHeight = startRect.height;

    // Adjust dimensions based on which edge/corner is being dragged
    if (direction.includes('e')) { newWidth  = startRect.width  + dx; }
    if (direction.includes('w')) { newWidth  = startRect.width  - dx; newLeft = startRect.left + dx; }
    if (direction.includes('s')) { newHeight = startRect.height + dy; }
    if (direction.includes('n')) { newHeight = startRect.height - dy; newTop  = startRect.top  + dy; }

    // Enforce minimum size
    if (newWidth  < MIN_W) { if (direction.includes('w')) newLeft = startRect.right - MIN_W; newWidth  = MIN_W; }
    if (newHeight < MIN_H) { if (direction.includes('n')) newTop  = startRect.bottom - MIN_H; newHeight = MIN_H; }

    // Clamp to viewport
    if (newTop  < 0) { newHeight += newTop; newTop = 0; }
    if (newLeft < 0) { newWidth  += newLeft; newLeft = 0; }
    if (newLeft + newWidth  > window.innerWidth)  newWidth  = window.innerWidth  - newLeft;
    if (newTop  + newHeight > window.innerHeight) newHeight = window.innerHeight - newTop;

    // Apply — switch to top/left/width/height positioning (clear right/bottom)
    panel.style.top    = newTop    + 'px';
    panel.style.left   = newLeft   + 'px';
    panel.style.width  = newWidth  + 'px';
    panel.style.height = newHeight + 'px';
    panel.style.right  = 'auto';
    panel.style.bottom = 'auto';
  }

  /** End the resize drag and restore normal cursor. */
  function onMouseUp() {
    if (!active) return;
    active = false;
    panel.classList.remove('resizing');
    document.body.style.cursor    = '';
    document.body.style.userSelect = '';
  }

  // Attach listeners to each resize handle
  panel.querySelectorAll('.ai-resize').forEach(h => h.addEventListener('mousedown', onMouseDown));
  document.addEventListener('mousemove', onMouseMove);
  document.addEventListener('mouseup', onMouseUp);
})();


/* ============================================================================
   WIRE UP AI ASSISTANT BUTTONS & BOOT
   ============================================================================ */

document.getElementById('ai-fab').addEventListener('click', aiOpen);
document.getElementById('ai-close').addEventListener('click', aiClose);
document.getElementById('ai-backdrop').addEventListener('click', aiClose);
document.getElementById('ai-send').addEventListener('click', aiSend);
document.getElementById('ai-input').addEventListener('keydown', (e) => { if (e.key === 'Enter') aiSend(); });

// Initial render of equipment data
render();

// Auto-open the assistant on first page load
aiOpen();
</script>
</body>
</html>"""


def generate_html(equipment, filter_options, output_path):
    from datetime import date
    data_json = json.dumps(equipment, separators=(',', ':'))
    filter_json = json.dumps(filter_options, separators=(',', ':'))
    html = HTML_TEMPLATE.replace('__DATA__', data_json)
    html = html.replace('__FILTER_OPTIONS__', filter_json)
    html = html.replace('__DATE__', date.today().strftime('%B %d, %Y'))
    html = html.replace('__ASSISTANT_PROXY_URL__', ASSISTANT_PROXY_URL)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  Written: {output_path}  ({len(html):,} bytes)")


if __name__ == '__main__':
    base = os.path.dirname(os.path.abspath(__file__))
    source = os.path.join(base, SOURCE_FILE)
    output = os.path.join(base, OUTPUT_FILE)

    if not os.path.exists(source):
        print(f"ERROR: Source file not found: {source}")
        print("Place this script in the same folder as the Excel file.")
        exit(1)

    print(f"Reading: {SOURCE_FILE}")
    equipment, filter_options = extract_data(source)

    print(f"Generating HTML…")
    generate_html(equipment, filter_options, output)

    print(f"\n✓ Done! Open {OUTPUT_FILE} in any browser.")
    print(f"  {len(equipment)} equipment entries across "
          f"{len(filter_options['departments'])} departments.")

    if ASSISTANT_PROXY_URL:
        print(f"\n  AI Equipment Assistant is wired to: {ASSISTANT_PROXY_URL}")
        print("    No API key is embedded in this HTML — confirm the Worker is")
        print("    deployed and ALLOWED_ORIGINS in worker.js includes the exact")
        print("    URL this page will be hosted at (e.g. your github.io URL).")
    else:
        print("\n  ℹ AI Equipment Assistant is disabled (ASSISTANT_PROXY_URL is empty).")
        print("    See DEPLOYMENT.md to deploy the free Cloudflare Worker proxy,")
        print("    then paste its URL into ASSISTANT_PROXY_URL near the top of")
        print("    this script and re-run. Do NOT put a Gemini API key directly")
        print("    in this script or in the generated HTML — it would be")
        print("    committed to your repo and exposed to every site visitor.")
